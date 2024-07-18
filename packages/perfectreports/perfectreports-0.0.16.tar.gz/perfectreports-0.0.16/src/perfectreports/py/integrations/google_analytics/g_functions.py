import openpyxl
import os
from base64 import b64encode
import plotly.graph_objects as go

def create_pie_analytics(df, title, column, value_column, style="height: 230px;width: 300px;"):
    layout = go.Layout(
        autosize=True,
        showlegend=False,
        title_text="<b>" + title + "</b>",
        title_x=0.5,
        title_y=0.98,
    )
    fig = go.Figure(data=[go.Pie(labels=df[column].tolist(),
                    values=df[value_column].values.tolist(), pull=[0.1, 0.05, 0.03, 0],
                    hole=.3, texttemplate="%{label}: %{value} <br>(%{percent})", insidetextorientation='horizontal', sort=True)], layout=layout)
    fig.update_traces(marker=dict(line=dict(color='white', width=3)))
    fig.update_layout(font=dict(family="sans-serif",
                      size=16, color="black"))
    img = fig.to_image(format="png", engine="kaleido")

    summary = '<img alt="execution summary" id="box" style="'+style+'" src="data:image/png;base64, {}" >'.format(
        b64encode(img).decode("utf-8"))
    return summary + df.to_html(
        table_id="table",
        index=False,
        render_links=False,
        escape=True,
    )


def getStartEndFromExcel(dirname, filename, key):
    dataframe = openpyxl.load_workbook(os.path.join(
        dirname, filename))

    dataframe1 = dataframe.active
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, 1):
            if col[row].value == key:
                start = row
                break
        else:
            continue
        break
    for row in range(start, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, 1):
            if col[row].value is None:
                end = row
                break
        else:
            continue
        break
    return start, end, dataframe1.max_row
