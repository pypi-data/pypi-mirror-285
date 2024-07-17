from loguru import logger
import pendulum
from pathlib import Path
from openpyxl import Workbook, load_workbook
from typing import List, Dict, Any, Tuple, Union

def write_list_tuple_excel(datas: Union[List[Tuple[Any]],List[List[Tuple[Any]]]], headers: List[str] = None, filename: str = None, sheet_name: Union[str,List,None]=None) -> None:
    """
    将一个或者多个列表元组写入到excel文件中
    datas:
        eg1. [('argo', 'elk'),('argo2', 'elk2')]
        eg2. [[('argo', 'elk'),('argo2', 'elk2')], [(...), (...)]]
    headers: 表头列表,默认None,不写表头; ['命名空间', '项目名称']
    filename: 写入excel中的文件名,默认为None
    sheet_name: 工作表名称,默认None;
        eg0: "hzUsa"
        eg1: 字符串传递: "杭州,美国"
        eg2: 列表传递: ["杭州","美国"]
    """
    if not isinstance(datas[0], (list, tuple)):
        raise ValueError("Data[0] must be a list or tuple.")

    # 判断文件是否已经存在,如果存在则删除
    if filename:
        if Path(filename).exists() or Path(f'{filename}.xlsx').exists():
            try:
                Path(filename).unlink()
                Path(f'{filename}.xlsx').unlink()
            except Exception:
                pass

    wb = Workbook()
    wb.encoding = 'utf-8'

    # 如果是多个list,则走下面的逻辑
    if isinstance(datas[0], list):
        for index, data in enumerate(datas, start=1):
            if sheet_name:
                try:
                    # 如果sheet_name是字符串
                    if isinstance(sheet_name, str):
                        ws = wb.create_sheet(title=sheet_name.split(',')[index-1])
                    if isinstance(sheet_name, list):
                        ws = wb.create_sheet(title=sheet_name[index-1])
                except Exception as e:
                    logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                    ws = wb.create_sheet(title=f"Sheet{index}")
            else:
                ws = wb.create_sheet(title=f"Sheet{index}")

            if headers:
                ws.append(headers)
            for item in data:
                ws.append(item)
            # 设置单元格宽度
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # 删除默认的Sheet
        wb.remove(wb['Sheet'])

    # 如果是单个list,则走下面的逻辑
    if isinstance(datas[0], tuple):
        if sheet_name:
            try:
                # 如果sheet_name是字符串
                if isinstance(sheet_name, str):
                    ws = wb.create_sheet(title=str(sheet_name))
                if isinstance(sheet_name, list):
                    ws = wb.create_sheet(title=sheet_name[0])
                # 删除默认的Sheet
                wb.remove(wb['Sheet'])
            except Exception as e:
                logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                ws = wb.active
        else:
            ws = wb.active

        if headers:
            ws.append(headers)
        for item in datas:
            ws.append(item)
        # 设置单元格宽度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 自动添加后缀xlsx
    if filename and not str(filename).endswith('.xlsx'):
        filename = filename +  '.xlsx'

    if not filename:
        filename = f'write-{pendulum.now().to_datetime_string()}.xlsx'

    wb.save(filename)
    logger.success(f'写入文件成功,文件名: {filename}')

def write_list_dict_excel(datas: Union[List[Dict[str, Any]],List[List[Dict[str, Any]]]], headers: Dict[str, str]=None, filename: str=None, sheet_name: Union[str,List,None]=None):
    """
    将一个或者多个字典列表写入excel中
    datas: eg1. [{'ns': 'argo', 'tp': 'elk'}, {'ns': 'argo2', 'tp': 'elk2'}]
           eg2. [[{'ns': 'argo', 'tp': 'elk'}, {'ns': 'argo2', 'tp': 'elk2'}], [{...}, {...}]]
    headers: 字典，每个键值对对应列名和标题；如果未提供，则根据第一个数据字典的键值自动生成; {'ns': '命名空间', 'tp': '类型'}
    filename: 响应头中的文件名,提供前端获取
    sheet_name: 工作表名称,默认None;
        eg0: "hzUsa"
        eg1: 字符串传递: "杭州,美国"
        eg2: 列表传递: ["杭州","美国"]
    """
    wb = Workbook()
    wb.encoding = 'utf-8'

    if not isinstance(datas[0], (list, dict)):
        raise ValueError("Data[0] must be a list or dict.")

    # 如果传递多个列表,则使用下面的创建逻辑
    if isinstance(datas[0], list):
        for index, data in enumerate(datas, start=1):
            if sheet_name:
                try:
                    # 如果sheet_name是字符串
                    if isinstance(sheet_name, str):
                        ws = wb.create_sheet(title=sheet_name.split(',')[index-1])
                    if isinstance(sheet_name, list):
                        ws = wb.create_sheet(title=sheet_name[index-1])
                except Exception as e:
                    logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                    ws = wb.create_sheet(title=f"Sheet{index}")
            else:
                ws = wb.create_sheet(title=f"Sheet{index}")

            # 判断data的第一个item是否为dict
            if not isinstance(data[0], dict):
                raise ValueError("Data[0] must be a dict.")

            # 获取所有字段名
            fieldnames = []
            for item in data:
                for k in item.keys():
                    if k not in fieldnames:
                        fieldnames.append(k)

            if headers and isinstance(headers, dict):
                header_values = [headers.get(key, key) for key in fieldnames]
                ws.append(header_values)
            else:
                ws.append(fieldnames)

           # 写入数据
            for item in data:
                ws.append([item.get(key) for key in fieldnames])

            # 设置单元格宽度
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # 删除默认的Sheet
        wb.remove(wb['Sheet'])

    # 如果是单个数据list,走下面的逻辑
    if isinstance(datas[0], dict):
        if sheet_name:
            try:
                # 如果sheet_name是字符串
                if isinstance(sheet_name, str):
                    ws = wb.create_sheet(title=str(sheet_name))
                if isinstance(sheet_name, list):
                    ws = wb.create_sheet(title=sheet_name[0])
                # 删除默认的Sheet
                wb.remove(wb['Sheet'])
            except Exception as e:
                logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                ws = wb.active
        else:
            ws = wb.active

        # 获取所有字段名
        fieldnames = []
        for item in datas:
            for k in item.keys():
                if k not in fieldnames:
                    fieldnames.append(k)

        if headers and isinstance(headers, dict):
            header_values = [headers.get(key, key) for key in list(fieldnames)]
            ws.append(header_values)
        else:
            ws.append(fieldnames)

        for item in datas:
            ws.append([item.get(key) for key in fieldnames])

        # 设置单元格宽度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 自动添加后缀xlsx
    if filename and not str(filename).endswith('.xlsx'):
        filename = filename +  '.xlsx'

    if not filename:
        filename = f'write-{pendulum.now().to_datetime_string()}.xlsx'

    wb.save(filename)
    logger.success(f'写入文件成功,文件名: {filename}')

def read_excel(file_path: str) -> dict:
    """
    读取excel文件,返回以sheetName为key的字典
    """
    # 判断文件是否存在并且要以.xlsx结尾
    if not Path(file_path).exists():
        raise ValueError(f"{file_path} 文件不存在")

    if not file_path.endswith('.xlsx'):
        raise ValueError(f"{file_path} 不是.xlsx文件")

    # 打开Excel文件
    workbook = load_workbook(filename=file_path)

    # 创建一个空字典来存储所有工作表的数据
    sheets_data = {}

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 读取工作表的所有内容
        data = []

        # 获取工作表的行数和列数
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 获取第一行作为字典的键，如果为空则使用默认名称
        headers = []
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is None:
                headers.append(f'col_{col}')
            else:
                headers.append(cell_value)

        # 从第二行开始读取数据，并将其存储为字典
        for row in range(2, max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data[header] = cell_value
            data.append(row_data)

        # 将当前工作表的数据存储到字典中
        sheets_data[sheet_name] = data

    return sheets_data

if __name__ == '__main__':
    file_path = 'write-2023-06-03 11:18:29.xlsx'  # 替换为你的Excel文件路径
    data = read_excel(file_path)
    # print(data)
    print(data['Sheet1'])
    print(data['Sheet2'])
    for item in data.values():
        print(item)

