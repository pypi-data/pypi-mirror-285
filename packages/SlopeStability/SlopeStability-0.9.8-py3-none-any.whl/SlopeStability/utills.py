import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import math 
from openpyxl import load_workbook
import pkg_resources
import os

def getDataFromExcel(file_name,sheet_name):
    # DATA_FOLDER = os.path.join(os.path.dirname(__file__), 'Data')
    
    # file_path = os.path.join(DATA_FOLDER, file_name)
    file_path = pkg_resources.resource_filename('SlopeStability', 'Data')
    file_path = os.path.join(file_path,file_name)
    # print(file_path)
    wb = load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2,values_only=True):
            # Filter out rows that are None or do not have at least 2 valid values
            if len(row) >= 2 and row[0] is not None and row[1] is not None:
                data.append([row[0], row[1]])  # Select only the first 2 columns

    # Convert the data to a numpy array
    data = np.array(data)
    return data
    pass

def find_closest_keys(d, keys):
    # This function finds the two keys closest to 'd' (one higher, one lower)
    arr = np.array(sorted(keys))
    low, high=-1,-1
    n = len(arr)
    
    for i in range(n-1):
        if d>arr[i] and d<arr[i+1]:
            low = (arr[i])
            high = (arr[i+1])
    if low==-1 and high==-1:
        raise ValueError("Point is out of bounds of the Stability Charts")
    return low,high
    pass

def linear_interpolate(x1, y1, x2, y2, x):
    # Perform linear interpolation
    if x2 - x1 == 0:  # To handle division by zero if x1 == x2
        return y1
    return y1 + (y2 - y1) * (x - x1) / (x2 - x1)
    pass

def interpolate_2d_array(points, x):
    # Ensure points are sorted by x values
    points = np.array(sorted(points, key=lambda point: point[0]))
    
    # Extract x and y arrays from the points
    x_points = points[:, 0]
    y_points = points[:, 1]
    
    # Check if x is out of the bounds of the provided points
    if x < x_points[0] or x > x_points[-1]:
        raise ValueError("Point is out of bounds of the Stability Charts.")
    
    # Find the indices of the two x points between which x lies
    for i in range(len(x_points) - 1):
        if x_points[i] <= x <= x_points[i + 1]:
            x0, y0 = x_points[i], y_points[i]
            x1, y1 = x_points[i + 1], y_points[i + 1]
            # Apply linear interpolation formula
            y = y0 + (x - x0) * (y1 - y0) / (x1 - x0)
            return y
    n = len(x_points)
    if math.isclose(x_points[n-1], x, rel_tol=1e-8, abs_tol=1e-8):
        return y_points[n-1]
    
    # If we can't find the interval, raise an error
    raise ValueError("Interpolation could not be performed. Check input data.")
    pass

def getFromDict(m,d,x):
    if d in m:
        result = interpolate_2d_array(m[d], x)
    else:
        # Find the closest keys for interpolation
        lower_key, upper_key = find_closest_keys(d, m.keys())
        # Calculate f for both lower and upper keys
        lower_value = interpolate_2d_array(m[lower_key], x)
        upper_value = interpolate_2d_array(m[upper_key], x)
        # Interpolate to find the result
        result = linear_interpolate(lower_key, lower_value, upper_key, upper_value, d)

    return result
    pass

def FailureCircle(x0,y0,D,H,beta,T,q,Hw):
    if D==0:
        D=H
    # Convert slope angle from degrees to radians
    beta_rad = np.radians(beta)

    # Create a figure and an axes
    fig, ax = plt.subplots(2,1,figsize=(8, 10))

    # Plot the slope line
    line_length = H / np.sin(beta_rad)
    line_x = [0, line_length * np.cos(beta_rad)]
    line_y = [0, H]

    # Ensure only positive x values (i.e., the line should point to the right)
    if line_x[1] < 0:
        line_x[1] = -line_x[1]
        line_y[1] = -line_y[1]

    # Plot the slope line
    ax[0].plot(line_x, line_y, label='Slope Line', color='black')
    
    a=1.5*H
    # Plot the horizontal line at the height of the slope
    x_values = np.linspace(H / np.tan(beta_rad), H / np.tan(beta_rad)+a, 40)  # Adjust max value if needed
    y_values = np.full_like(x_values, H)  # y=H for all x in x_values
    ax[0].plot(x_values, y_values, color='black', label='Horizontal Line at H')


# Plot the horizontal line at the depth of the foundation
    x_values_depth = np.linspace(-(a), 0, 40)
    y_values_depth = np.full_like(x_values_depth, 0)  # y=0 for all x in x_values_depth
    ax[0].plot(x_values_depth, y_values_depth, color='black', label='Foundation Depth Line')
    
    x_values_depth = np.linspace(-(a),H/np.tan(beta_rad)+a , 80)
    y_values_depth = np.full_like(x_values_depth, -D)  # y=0 for all x in x_values_depth
    ax[0].plot(x_values_depth, y_values_depth, color='black', label='Foundation Depth Line')
    
    x_values_depth = np.linspace(0, H/np.tan(beta_rad)+a, 40)
    y_values_depth = np.full_like(x_values_depth,0)  # y=0 for all x in x_values_depth
    ax[0].plot(x_values_depth, y_values_depth, linestyle='--',color='black', label='Foundation Depth Line')

    x_const = -a
    y_range = np.linspace(0,-D, 40)
    # Plot the line parallel to y-axis
    ax[0].plot([x_const]*len(y_range), y_range,color='black')
    
    x_const = H/np.tan(beta_rad)+a
    y_range = np.linspace(-D,H, 40)
    # Plot the line parallel to y-axis
    ax[0].plot([x_const]*len(y_range), y_range,color='black')
    
    if q!=0:
        num_arrows = 5  # Number of arrows
        x_arrows = np.linspace(H / np.tan(beta_rad), H/np.tan(beta_rad)+a, num_arrows)

        # Length of each arrow
        arrow_length = 3  # Shorter length
        i=0
        for x in x_arrows:
            if(i==np.round(len(x_arrows)/2)):
                ax[0].annotate(f'q={q}', xy=(x, H), xytext=(x,H+arrow_length),arrowprops=dict(facecolor='red', shrink=1))  # Adjust shrink for smaller arrow
            else:
                ax[0].annotate('', xy=(x, H), xytext=(x,H+arrow_length),arrowprops=dict(facecolor='red', shrink=1))
            i=i+1
    if Hw!=0:
        x_values_depth = np.linspace(-(a), Hw/np.tan(beta_rad), 40)
        y_values_depth = np.full_like(x_values_depth, Hw)  # y=0 for all x in x_values_depth
        ax[0].plot(x_values_depth, y_values_depth, color='blue', label='Foundation Depth Line')
        x_const = -a
        y_range = np.linspace(0,Hw, 40)
        # Plot the line parallel to y-axis
        ax[0].plot([x_const]*len(y_range), y_range,color='blue')
        ax[0].annotate('', xy=(-a/2, Hw), xytext=(-a/2,Hw+2),arrowprops=dict(facecolor='blue', shrink=0.1))
        ax[0].annotate(f'Hw: {Hw}', xy=(-a, Hw/2), xytext=(-(a+9), Hw/2),
                arrowprops=dict(facecolor='red', arrowstyle='->'), fontsize=12, color='red')
        
    # Annotate the slope angle
    ax[0].annotate(f'Angle: {beta}°', xy=(line_x[1]/2, line_y[1]/2), xytext=(line_x[1]/2 + 5, line_y[1]/2),
                arrowprops=dict(facecolor='red', arrowstyle='->'), fontsize=12, color='red')
    if D!=0:
        ax[0].annotate(f'D: {D}', xy=(-a, -D/2), xytext=(-(a+9), -D/2),
                arrowprops=dict(facecolor='red', arrowstyle='->'), fontsize=12, color='red')
        ax[0].annotate(f'D: {D}', xy=(H/np.tan(beta_rad)+a, -D/2), xytext=(H/np.tan(beta_rad)+a+5, -D/2),
            arrowprops=dict(facecolor='red', arrowstyle='->'), fontsize=12, color='red')

    ax[0].annotate(f'H: {H}', xy=(H/np.tan(beta_rad)+a, H/2), xytext=(H/np.tan(beta_rad)+a+5, H/2),
                arrowprops=dict(facecolor='red', arrowstyle='->'), fontsize=12, color='red')
    
    # Set the aspect of the plot to be equal
    ax[0].set_aspect('equal')

    # Set labels and title
    ax[0].set_xlabel('x')
    ax[0].set_ylabel('y')
    ax[0].set_title('Problem statement',pad=20)
    ax[0].axis('off')
    
    # Define the circle's center and radius
    center = (x0, y0)
    if T==1:
        radius = np.sqrt(x0*x0+y0*y0)
    else:
        radius = y0+D
    beta=beta_rad

    # Create an array of angles from 0 to 2pi
    theta = np.linspace(0, 2 * np.pi, 2000)

    # Parametric equations for the circle
    x = center[0] + radius * np.cos(theta)
    y = center[1] + radius * np.sin(theta)

    # Filter the circle points to show only where y <= 5 and x >= 0
    x1 = []
    y1 = []

    # Iterate through the arrays and apply the conditions
    for i in range(len(x)):
        if (x[i] >= 0 and y[i] <= H) or (y[i]<=0 and x[i]<=0) :
            x1.append(x[i])
            y1.append(y[i])
    # Plot the filtered circle
    ax[1].plot(x1, y1, label='Filtered Circle: y<=H and x>=0')


    # Define the line properties
    line_length = H/np.sin(beta)
    line_angle_rad = beta

    # Calculate the endpoints of the line
    line_x = [0, line_length * np.cos(line_angle_rad)]
    line_y = [0, line_length * np.sin(line_angle_rad)]

    # Ensure only positive x values (i.e., the line should point to the right)
    if line_x[1] < 0:
        line_x[1] = -line_x[1]
        line_y[1] = -line_y[1]

    # Plot the line
    ax[1].plot(line_x, line_y, label='Line from origin, length H', color='black')

    x_values = np.linspace(H/np.tan(beta), max(40, 2*x0+radius), 40)  # Adjust max value if needed
    y_values = np.full_like(x_values, H)  # y=H for all x in x_values
    ax[1].plot(x_values, y_values, color='black')
    
    x_values_y0 = np.linspace(-np.sqrt(radius*radius-y0*y0), 0, 100)  # Adjust min value if needed
    y_values_y0 = np.full_like(x_values_y0, 0)  # y=0 for all x in x_values_y0
    ax[1].plot(x_values_y0, y_values_y0, color='black', label='Line y=0 for x<=0')
    
    x_values_y0 = np.linspace(-np.sqrt(radius*radius-y0*y0), max(40, 2*x0+radius), 100)  # Adjust min value if needed
    y_values_y0 = np.full_like(x_values_y0, -D)  # y=0 for all x in x_values_y0
    ax[1].plot(x_values_y0, y_values_y0, color='black', label='Line y=0 for x<=0')
    
    x_const = max(40, 2*x0+radius)
    y_range = np.linspace(-D,H, 40)
    # Plot the line parallel to y-axis
    ax[1].plot([x_const]*len(y_range), y_range,color='black')
    
    x_const = -np.sqrt(radius*radius-y0*y0)
    y_range = np.linspace(-D,0, 40)
    # Plot the line parallel to y-axis
    ax[1].plot([x_const]*len(y_range), y_range,color='black')

    # Set the aspect of the plot to be equal
    ax[1].set_aspect('equal')

    # Set labels and title
    ax[1].set_xlabel('x')
    ax[1].set_ylabel('y')
    if T==1:
        str='Toe circle'
    elif T==2:
        str='Deep circle'
    elif T==3:
        str='Slope circle'
    ax[1].set_title(str)
    ax[1].axis('off')

    # Show the plot
    plt.show()