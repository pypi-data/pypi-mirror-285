import os
import requests
import pandas as pd
from io import StringIO
import base64
from io import StringIO
from datetime import datetime
from openpyxl import load_workbook

delta = 2654435769

def join(numbers):
    bytes_array = bytearray()
    for number in numbers:
        bytes_array.extend(number.to_bytes(4, "little", signed=False))
    return bytes_array

def split(bytes_array):
    numbers = []
    bytes_array_length = len(bytes_array)
    numbers_length = bytes_array_length // 4

    if bytes_array_length % 4 > 0:
        numbers_length += 1
        bytes_array.extend(bytearray(4 - bytes_array_length % 4))

    index = 0
    while index < numbers_length:
        number = int.from_bytes(
            bytes_array[index * 4 : (index + 1) * 4], "little", signed=False
        )
        numbers.append(number)
        index += 1

    return numbers

def run(text_bytes, key_bytes):
    text_numbers = split(text_bytes)
    key_numbers = split(key_bytes)
    if len(key_numbers) < 4:
        key_numbers.extend([0] * (4 - len(key_numbers)))

    text_length = len(text_numbers) - 1
    text_first = text_numbers[0]
    rounds = (6 + 52 // (text_length + 1)) * delta & 0xFFFFFFFF

    while rounds != 0:
        q = (rounds >> 2) & 3
        index = text_length

        while index > 0:
            current = text_numbers[index - 1]
            p = (current >> 5 ^ text_first << 2) + (text_first >> 3 ^ current << 4) ^ (
                rounds ^ text_first
            ) + (key_numbers[index & 3 ^ q] ^ current)
            temp = (text_numbers[index] - p) & 0xFFFFFFFF
            if temp < 0:
                temp += 2**32
            text_first = text_numbers[index] = temp
            index -= 1

        current = text_numbers[text_length]
        p = (current >> 5 ^ text_first << 2) + (text_first >> 3 ^ current << 4) ^ (
            rounds ^ text_first
        ) + (key_numbers[index & 3 ^ q] ^ current)
        temp = (text_numbers[0] - p) & 0xFFFFFFFF
        if temp < 0:
            temp += 2**32

        text_first = text_numbers[0] = temp
        rounds -= delta

        if rounds < 0:
            rounds += 2**32

    return join(text_numbers)

def decode(raw_text):
    suffix = "03a33cd9a31ee58c"
    key = "redwar2021"
    text = raw_text[: -len(suffix)]
    text_bytes = base64.b64decode(text)
    key_bytes = bytearray(key.encode("utf-8"))
    result = run(text_bytes, key_bytes)
    return result.decode("utf-8", errors="replace")

def fetch_tsv_data(url):
    try:
        print("trying: " + url[-14:-4])
        response = requests.get(url)
        response.raise_for_status()
        response.encoding = "gbk"
        return decode(response.text)
    except requests.RequestException:
        return None

def get_act_time():
    '''
    尝试获取活动时间excel,保存至桌面
    '''
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    output_file = os.path.join(desktop_path, "temp.xlsx")

    data_frames = []
    current_date = int(datetime.now().strftime(r"%Y%m%d"))
    for i in range(current_date,20240701, -1):
        if i%100 >31:
            continue
        success=False
        for j in range(3,0, -1):
            url = f"https://100616028cdn-1251006671.file.myqcloud.com/100616028/res/20120522/config/GameConstantSet_{i:08d}{j:02d}.dat"
            tsv_data = fetch_tsv_data(url)
            if tsv_data:
                print("success：" + url)
                tsv_io = StringIO(tsv_data)
                df = pd.read_csv(tsv_io, sep="\t")
                data_frames.append(df)
                success=True
                output_file=output_file.replace('temp',f'{i:08d}{j:02d}')
                break
        if success:
            break

    if data_frames:
        combined_df = pd.concat(data_frames, ignore_index=True)
        filtered_df = combined_df[combined_df.iloc[:, 4] == '开始时间DAY']
        filtered_df.iloc[:, 3] = pd.to_datetime(filtered_df.iloc[:, 3])
        sorted_df = filtered_df.sort_values(by=filtered_df.columns[3], ascending=False)
        columns_to_drop = [0,2,4,5,6,8,9,10]
        sorted_df.drop(sorted_df.columns[columns_to_drop], axis=1, inplace=True)
        sorted_df.to_excel(output_file, index=False)
        print(f"file: {output_file}")
        wb = load_workbook(output_file)
        ws = wb.active
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        date_format = "yyyy-mm-dd"  
        for cell in ws['B']:  
            cell.number_format = date_format
        for cell in ws['C']:  
            cell.number_format = date_format
        wb.save(output_file)
    else:
        print("filed!")
