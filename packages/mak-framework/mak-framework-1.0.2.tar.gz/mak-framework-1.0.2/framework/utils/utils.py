import csv
import os
import json
import random
import string
import shutil
import allure
import logging
import platform
import requests
import subprocess
import pandas as pd
from allure_commons.types import LinkType
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook
from logging.handlers import RotatingFileHandler
from framework.utils.create_general_directory import RESULTS_DIR


class TestResultManager:
    @staticmethod
    def get_test_result_folder_name():
        folder_name = (os.environ.get('PYTEST_CURRENT_TEST')
                       .replace('::::', '-')
                       .replace('.', '-')
                       .replace('_', '-')
                       .replace(' (setup)', '')
                       .replace(' (call)', '')
                       .lower())
        return folder_name

    @staticmethod
    def get_test_output_dir():
        return os.path.join(
            os.path.join(f"{RESULTS_DIR}\\Test_Results\\", TestResultManager.get_test_result_folder_name()))

    @staticmethod
    def write_file(directory, name_with_extension, content):
        folder_path = directory
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        file_name = name_with_extension
        file_path = os.path.join(folder_path, file_name)
        with open(file_path, 'w') as f:
            f.write(content)

    @staticmethod
    def attach_artifact_folder_link_to_allure():
        TestResultManager.write_file(TestResultManager.get_test_output_dir(), "execution.bat",
                                     f"playwright show-trace trace.zip\nexit")
        TestResultManager.write_file(RESULTS_DIR, "allure_single_file.bat",
                                     "allure generate --single-file Allure_Results\nexit")
        TestResultManager.write_file(RESULTS_DIR, "allure_serve.bat", "allure serve Allure_Results")
        allure.dynamic.link(url=TestResultManager.get_test_output_dir(), link_type=LinkType.TEST_CASE,
                            name=TestResultManager.get_test_output_dir())


class FileDataManager:
    @staticmethod
    def get_test_data(file_name, file_extension=".json"):
        """
        Get test data from a JSON file.

        Args:
            file_name (str): Name of the JSON file.
            file_extension (str, optional): Extension of the file. Default is ".json".

        Returns:
            dict: Test data read from the JSON file.
        """
        file_path = FileDataManager.get_file_location(file_name, file_extension)
        return JSONUtils.read_json(file_path)

    @staticmethod
    def get_data_from_file(file_path, sheet_name='Sheet1'):
        complete_file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "R2_Data_Config", file_path)
        if file_path.endswith('.csv'):
            df = pd.read_csv(complete_file_path)
        elif file_path.endswith('.xlsx', '.xls'):
            df = pd.read_excel(complete_file_path, sheet_name=sheet_name)
        else:
            raise ValueError("Unsupported file format. Please provide CSV or Excel file.")
        data = df.itertuples(index=False, name=None)
        return list(data)

    @staticmethod
    def get_file_location(filename, file_extension=""):
        current_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        print(f"Searching for file: {filename} with extension: {file_extension} in directory: {current_dir}")

        for root, dirs, files in os.walk(current_dir):
            for file in files:
                if file_extension == "" or file.endswith(file_extension):
                    if file == filename:
                        file_location = os.path.join(root, file)
                        print(f"Found file at location: {file_location}")
                        return file_location

        print(f"File not found: {filename} with extension: {file_extension}")
        return None


class ExcelUtils:
    @staticmethod
    def read_excel(file_path, sheet_name=0):
        return pd.read_excel(file_path, sheet_name=sheet_name)

    @staticmethod
    def write_excel(file_path, data, sheet_name='Sheet1'):
        df = pd.DataFrame(data)
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

    @staticmethod
    def append_to_excel(file_path, data, sheet_name='Sheet1'):
        df = pd.DataFrame(data)
        try:
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if sheet_name in book.sheetnames:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    combined_df = pd.concat([existing_df, df])
                    del book[sheet_name]
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        except FileNotFoundError:
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
        except Exception as e:
            raise e


class CSVUtils:
    @staticmethod
    def read_csv(file_path):
        with open(file_path, mode='r') as file:
            reader = csv.DictReader(file)
            return [row for row in reader]

    @staticmethod
    def write_csv(file_path, data):
        with open(file_path, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def append_to_csv(file_path, data):
        with open(file_path, mode='a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=data[0].keys())
            writer.writerows(data)


class JSONUtils:
    @staticmethod
    def read_json(file_path):
        with open(file_path, 'r') as file:
            return json.load(file)

    @staticmethod
    def read_json_file(filename):
        with open(filename, "r") as file:
            try:
                json_data = json.load(file)
                return json_data
            except json.JSONDecodeError as e:
                raise json.JSONDecodeError(f"Error: Invalid JSON format in file '{filename}': {e}")

    @staticmethod
    def write_json(file_path, data):
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)

    @staticmethod
    def update_json(file_path, key, value):
        data = JSONUtils.read_json(file_path)
        data[key] = value
        JSONUtils.write_json(file_path, data)

    @staticmethod
    def delete_key(file_path, key):
        data = JSONUtils.read_json(file_path)
        if key in data:
            del data[key]
            JSONUtils.write_json(file_path, data)

    @staticmethod
    def key_exists(file_path, key):
        data = JSONUtils.read_json(file_path)
        return key in data

    @staticmethod
    def get_value(file_path, key):
        data = JSONUtils.read_json(file_path)
        return data.get(key, None)


class WaitUtils:
    @staticmethod
    def wait_for_element(page, selector: str, timeout: int = 30) -> bool:
        try:
            page.wait_for_selector(selector, timeout=timeout * 1000)
            return True
        except:
            return False

    @staticmethod
    def wait_for_page_load(page, timeout: int = 30) -> bool:
        try:
            page.wait_for_load_state('load', timeout=timeout * 1000)
            return True
        except:
            return False


class RandomUtils:
    @staticmethod
    def generate_random_string(length=10):
        characters = string.ascii_letters + string.digits + string.punctuation
        return ''.join(random.choice(characters) for _ in range(length))

    @staticmethod
    def generate_random_cnic():
        five = str(random.randint(10000, 99999))
        seven = str(random.randint(1000000, 9999999))
        one = str(random.randint(0, 9))
        random_cnic = f"{five}-{seven}-{one}"
        return random_cnic

    @staticmethod
    def generate_random_number(minimum, maximum):
        return random.randint(minimum, maximum)

    @staticmethod
    def get_current_time_and_date(format_str="%Y-%m-%d %H:%M:%S"):
        current_time_and_date = datetime.now()
        formatted_date = current_time_and_date.strftime(format_str)
        return formatted_date

    @staticmethod
    def get_current_date(format_str="%d/%m/%Y", increment=None):
        current_date_and_time = datetime.now()
        if increment is not None:
            current_date_and_time += timedelta(days=increment)
        formatted_date = current_date_and_time.strftime(format_str)
        return formatted_date

    @staticmethod
    def generate_digit_random_number(minimum, maximum, num_digits):
        random_number_generator = ''
        while len(random_number_generator) < num_digits:
            digit = str(random.randint(minimum, maximum))
            random_number_generator += digit
        random_number_generator = random_number_generator[:num_digits]
        return int(random_number_generator)

    @staticmethod
    def generate_random_string1(length=10):
        letters = string.ascii_letters
        return ''.join(random.choice(letters) for _ in range(length))

    @staticmethod
    def generate_random_number1():
        return random.randint(1, 100)

    @staticmethod
    def generate_random_email(domain='example.com'):
        return f'{RandomUtils.generate_random_number1()}@{domain}'


class FileUtils:
    @staticmethod
    def create_directory(path):
        if not os.path.exists(path):
            os.makedirs(path)

    @staticmethod
    def delete_directory(path):
        if os.path.exists(path):
            shutil.rmtree(path)

    @staticmethod
    def copy_file(src, dest):
        shutil.copy(src, dest)

    @staticmethod
    def move_file(src, dest):
        shutil.move(src, dest)

    @staticmethod
    def delete_file(file_path):
        if os.path.exists(file_path):
            os.remove(file_path)

    @staticmethod
    def list_files(directory):
        """
        List all files in a directory (excluding subdirectories).
        """
        files = []
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            if os.path.isfile(file_path):
                files.append(filename)
        return files

    @staticmethod
    def list_directories(directory):
        """
        List all directories (excluding files) in a directory.
        """
        directories = []
        for filename in os.listdir(directory):
            dir_path = os.path.join(directory, filename)
            if os.path.isdir(dir_path):
                directories.append(filename)
        return directories

    @staticmethod
    def move_files_in_directory(src_dir, dest_dir):
        """
        Move all files from one directory to another.
        """
        FileUtils.create_directory(dest_dir)
        for filename in os.listdir(src_dir):
            src_path = os.path.join(src_dir, filename)
            dest_path = os.path.join(dest_dir, filename)
            shutil.move(src_path, dest_path)

    @staticmethod
    def copy_directory(src_dir, dest_dir):
        """
        Recursively copy a directory and its contents to another location.
        """
        shutil.copytree(src_dir, dest_dir)

    @staticmethod
    def rename_file(old_name, new_name):
        """
        Rename a file.
        """
        if os.path.exists(old_name):
            os.rename(old_name, new_name)
        else:
            raise FileNotFoundError(f"File '{old_name}' not found.")

    @staticmethod
    def change_file_permissions(file_path, mode):
        """
        Change the permissions (mode) of a file.
        """
        if os.path.exists(file_path):
            os.chmod(file_path, mode)
        else:
            raise FileNotFoundError(f"File '{file_path}' not found.")


class NetworkUtils:
    @staticmethod
    def get(url):
        response = requests.get(url)
        return response

    @staticmethod
    def post(url, data):
        response = requests.post(url, data=data)
        return response

    @staticmethod
    def delete(url):
        response = requests.delete(url)
        return response

    @staticmethod
    def put(url, data):
        response = requests.put(url, data=data)
        return response


class OSUtils:
    @staticmethod
    def get_os_name():
        return platform.system()

    @staticmethod
    def get_os_version():
        return platform.version()

    @staticmethod
    def list_directory(path='.'):
        try:
            return os.listdir(path)
        except FileNotFoundError:
            return []

    @staticmethod
    def create_directory(path):
        if not os.path.exists(path):
            os.makedirs(path)

    @staticmethod
    def delete_directory(path):
        if os.path.exists(path):
            shutil.rmtree(path)

    @staticmethod
    def copy_file(src, dest):
        shutil.copy(src, dest)

    @staticmethod
    def move_file(src, dest):
        shutil.move(src, dest)

    @staticmethod
    def delete_file(file_path):
        if os.path.exists(file_path):
            os.remove(file_path)

    @staticmethod
    def get_file_size(file_path):
        if os.path.exists(file_path):
            return os.path.getsize(file_path)
        return 0

    @staticmethod
    def execute_command(command):
        result = subprocess.run(command, shell=True, capture_output=True, text=True)
        return result.stdout.strip(), result.stderr.strip(), result.returncode

    @staticmethod
    def set_environment_variable(key, value):
        os.environ[key] = value

    @staticmethod
    def get_environment_variable(key, default=None):
        return os.getenv(key, default)

    @staticmethod
    def unset_environment_variable(key):
        if key in os.environ:
            del os.environ[key]


class DateTimeUtils:
    @staticmethod
    def get_current_datetime():
        return datetime.now()

    @staticmethod
    def add_days_to_current_date(days):
        return datetime.now() + timedelta(days=days)

    @staticmethod
    def format_datetime(dt, format='%d-%m-%Y__%H-%M-%S'):
        return dt.strftime(format)

    @staticmethod
    def test_get_current_datetime():
        current_datetime = DateTimeUtils.get_current_datetime()
        assert current_datetime.year == 2024

    @staticmethod
    def test_add_days_to_current_date():
        future_date = DateTimeUtils.add_days_to_current_date(7)
        assert future_date > DateTimeUtils.get_current_datetime()


class AllureUtils:
    @staticmethod
    def _get_timestamp():
        return datetime.now().strftime("%d-%m-%Y__%H-%M-%S")

    @staticmethod
    def attach_screenshot(driver, name='screenshot'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(driver.screenshot_as_png, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.PNG)

    @staticmethod
    def attach_text(text, name='text_attachment', attachment_type=allure.attachment_type.TEXT):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(text, name=f"{name}_{timestamp}", attachment_type=attachment_type)

    @staticmethod
    def attach_html(html_content, name='html_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(html_content, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.HTML)

    @staticmethod
    def attach_file(file_path, name=None):
        timestamp = AllureUtils._get_timestamp()
        allure.attach.file(file_path, name=f"{name or os.path.basename(file_path)}_{timestamp}")

    @staticmethod
    def attach_json(json_data, name='json_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(json.dumps(json_data), name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.JSON)

    @staticmethod
    def attach_xml(xml_data, name='xml_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(xml_data, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.XML)

    @staticmethod
    def attach_pdf(pdf_data, name='pdf_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(pdf_data, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.PDF)

    @staticmethod
    def attach_jpg(jpg_data, name='jpg_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(jpg_data, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.JPG)

    @staticmethod
    def attach_gif(gif_data, name='gif_attachment'):
        timestamp = AllureUtils._get_timestamp()
        allure.attach(gif_data, name=f"{name}_{timestamp}", attachment_type=allure.attachment_type.GIF)

    @staticmethod
    def attach_screenshot_to_allure(self, step_name: str = "Screenshot"):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        screenshots_dir = os.path.join(RESULTS_DIR, "Screenshots")
        os.makedirs(screenshots_dir, exist_ok=True)
        screenshot_path = os.path.join(screenshots_dir, f"{step_name.replace(' ', '_').lower()}_{timestamp}.png")
        self.page.screenshot(path=screenshot_path)
        allure.attach.file(screenshot_path, attachment_type=allure.attachment_type.PNG)

    @staticmethod
    def attach_video_to_allure(self, step_name: str = "Complete Test Case Video"):
        with allure.step(step_name):
            allure.attach.file(self.page.video.path(), attachment_type=allure.attachment_type.WEBM)


class LoggerUtils:
    LOG_LEVEL = logging.INFO
    LOG_FILE = "execution.log"
    LOG_FORMAT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    MAX_BYTES = 10485760
    BACKUP_COUNT = 3

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(self.LOG_LEVEL)
        formatter = logging.Formatter(self.LOG_FORMAT)

        if not self.logger.hasHandlers():
            self._add_stream_handler(formatter)
            self._add_file_handler(formatter)

    def _add_stream_handler(self, formatter):
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(self.LOG_LEVEL)
        stream_handler.setFormatter(formatter)
        self.logger.addHandler(stream_handler)

    def _add_file_handler(self, formatter):
        file_handler = RotatingFileHandler(self.LOG_FILE, maxBytes=self.MAX_BYTES, backupCount=self.BACKUP_COUNT)
        file_handler.setLevel(self.LOG_LEVEL)
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

    def log_info(self, message):
        self.logger.info(message)

    def log_debug(self, message):
        self.logger.debug(message)

    def log_warning(self, message):
        self.logger.warning(message)

    def log_error(self, message):
        self.logger.error(message)

    def log_critical(self, message):
        self.logger.critical(message)
